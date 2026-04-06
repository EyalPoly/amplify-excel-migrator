from typing import Dict, List, Any, Optional, Set, Tuple
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

_HEADER_FILL = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_HEADER_FONT = Font(bold=True)

_AMPLIFY_GENERATED_SUFFIXES = ("Connection", "Input", "FilterInput", "ConditionInput", "KeyInput")
_AMPLIFY_GENERATED_PREFIXES = ("__", "Model", "Searchable")
_AMPLIFY_SYSTEM_TYPES = {"Query", "Mutation", "Subscription"}


def _write_header(ws, columns: List[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _auto_size_columns(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)


class SchemaExporter:
    def __init__(self, client, field_parser):
        self.client = client
        self.field_parser = field_parser

    def export_to_markdown(self, output_path: str, models: Optional[List[str]] = None) -> None:
        logger.info(f"Exporting schema to {output_path}")

        if models is None:
            models = self.discover_models()

        markdown_content = self._generate_markdown(models)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write(markdown_content)

        logger.info(f"Schema exported successfully to {output_path}")

    def export_to_excel(self, output_path: str, models: Optional[List[str]] = None) -> None:
        logger.info(f"Exporting schema to {output_path}")

        if models is None:
            models = self.discover_models()

        # Fetch enums and custom types directly from the schema
        enums = self.client.get_all_enums()
        custom_type_names = self.discover_custom_types()
        custom_types = {name: fields for name in custom_type_names if (fields := self._get_custom_type_fields(name))}

        # Collect model rows
        model_data: List[Tuple[str, List[Dict[str, Any]]]] = []
        for model_name in models:
            logger.info(f"Processing model: {model_name}")
            rows = self._parse_model_fields(model_name)
            if rows is not None:
                model_data.append((model_name, rows))

        # Compute friendly display names
        known_prefixes: Set[str] = set(models) | set(custom_types.keys())
        display_names = self._get_enum_display_names(enums, known_prefixes)
        display_enums = self._build_display_enums(enums, display_names)

        # Write workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for model_name, rows in model_data:
            ws = wb.create_sheet(title=self._truncate_sheet_name(model_name))
            _write_header(ws, ["Field Name", "Type", "Required", "Description"])
            for row in rows:
                ws.append(
                    [
                        row["field_name"],
                        self._apply_display_names(row["type_display"], display_names),
                        "✅" if row["is_required"] else "❌",
                        row["description"] or "",
                    ]
                )
            _auto_size_columns(ws)

        if display_enums:
            ws = wb.create_sheet(title="Enums")
            _write_header(ws, ["Enum Name", "Allowed Values"])
            for enum_name, values in sorted(display_enums.items()):
                ws.append([enum_name, ", ".join(values)])
            _auto_size_columns(ws)

        if custom_types:
            ws = wb.create_sheet(title="Custom Types")
            _write_header(ws, ["Type Name", "Field Name", "Type", "Required"])
            for type_name, fields in sorted(custom_types.items()):
                for f in fields:
                    type_display = self._apply_display_names(self._format_type_display(f), display_names)
                    ws.append([type_name, f["name"], type_display, "✅" if f["is_required"] else "❌"])
            _auto_size_columns(ws)

        wb.save(output_path)
        logger.info(f"Schema exported successfully to {output_path}")

    def discover_models(self) -> List[str]:
        logger.info("Discovering models from schema")
        all_types = self.client.get_all_types()

        if not all_types:
            logger.warning("Could not introspect schema types")
            return []

        query_structure = self.client.get_model_structure("Query")
        query_field_names = set()
        if query_structure and "fields" in query_structure:
            query_field_names = {field.get("name", "") for field in query_structure["fields"]}

        models = set()
        for type_info in all_types:
            type_name = type_info.get("name", "")
            type_kind = type_info.get("kind", "")

            if type_kind == "OBJECT" and self._is_user_defined_type(type_name):
                has_list_query = any(
                    field_name.startswith("list") and type_name.lower() in field_name.lower() and "By" not in field_name
                    for field_name in query_field_names
                )
                if has_list_query:
                    models.add(type_name)

        return sorted(list(models))

    def discover_custom_types(self) -> List[str]:
        """Return OBJECT types that are custom types (not models, not Amplify-generated infra types)."""
        all_types = self.client.get_all_types()

        if not all_types:
            return []

        query_structure = self.client.get_model_structure("Query")
        query_field_names = set()
        if query_structure and "fields" in query_structure:
            query_field_names = {field.get("name", "") for field in query_structure["fields"]}

        custom_types = []
        for type_info in all_types:
            type_name = type_info.get("name", "")
            type_kind = type_info.get("kind", "")

            if type_kind != "OBJECT" or not self._is_user_defined_type(type_name):
                continue

            has_list_query = any(
                field_name.startswith("list") and type_name.lower() in field_name.lower() and "By" not in field_name
                for field_name in query_field_names
            )
            if not has_list_query:
                custom_types.append(type_name)

        return sorted(custom_types)

    def _generate_markdown(self, models: List[str]) -> str:
        # Fetch enums and custom types directly from the schema
        enums = self.client.get_all_enums()
        custom_type_names = self.discover_custom_types()
        custom_types = {name: fields for name in custom_type_names if (fields := self._get_custom_type_fields(name))}

        # Collect model sections
        model_sections: List[List[str]] = []
        for model in models:
            logger.info(f"Processing model: {model}")
            section = self._generate_model_section(model)
            if section:
                model_sections.append(section)

        # Compute display names
        known_prefixes: Set[str] = set(models) | set(custom_types.keys())
        display_names = self._get_enum_display_names(enums, known_prefixes)
        display_enums = self._build_display_enums(enums, display_names)

        lines = [
            "# GraphQL Schema Reference",
            "",
            "This document provides a complete reference of your GraphQL schema for preparing Excel data migrations.",
            "",
            "## Table of Contents",
            "",
        ]
        for model in models:
            lines.append(f"- [{model}](#{model.lower()})")
        lines.append("")
        lines.append("---")
        lines.append("")

        for section in model_sections:
            lines.extend(self._apply_display_names(line, display_names) for line in section)

        if display_enums:
            lines.append("---")
            lines.append("")
            lines.append("## Enums")
            lines.append("")
            for enum_name, enum_values in sorted(display_enums.items()):
                lines.append(f"### {enum_name}")
                lines.append("")
                for value in enum_values:
                    lines.append(f"- `{value}`")
                lines.append("")

        if custom_types:
            lines.append("---")
            lines.append("")
            lines.append("## Custom Types")
            lines.append("")
            for type_name, type_fields in sorted(custom_types.items()):
                lines.append(f"### {type_name}")
                lines.append("")
                lines.append("| Field Name | Type | Required |")
                lines.append("|------------|------|----------|")
                for field in type_fields:
                    required = "✅ Yes" if field["is_required"] else "❌ No"
                    type_display = self._apply_display_names(self._format_type_display(field), display_names)
                    lines.append(f"| {field['name']} | {type_display} | {required} |")
                lines.append("")

        return "\n".join(lines)

    def _generate_model_section(self, model_name: str) -> Optional[List[str]]:
        rows = self._parse_model_fields(model_name)
        if rows is None:
            return None

        raw_structure = self.client.get_model_structure(model_name)
        parsed_model = self.field_parser.parse_model_structure(raw_structure)

        lines = [f"## {model_name}", ""]

        if parsed_model.get("description"):
            lines.append(parsed_model["description"])
            lines.append("")

        lines.append("**Excel Sheet Name:** `" + model_name + "`")
        lines.append("")

        if not rows:
            lines.append("*No user-definable fields*")
            lines.append("")
            return lines

        lines.append("| Field Name | Type | Required | Description |")
        lines.append("|------------|------|----------|-------------|")

        for row in rows:
            required = "✅ Yes" if row["is_required"] else "❌ No"
            lines.append(f"| {row['field_name']} | {row['type_display']} | {required} | {row['description']} |")

        lines.append("")
        lines.append("---")
        lines.append("")

        return lines

    def _parse_model_fields(self, model_name: str) -> Optional[List[Dict[str, Any]]]:
        raw_structure = self.client.get_model_structure(model_name)
        if not raw_structure:
            logger.warning(f"Could not get structure for model: {model_name}")
            return None

        parsed_model = self.field_parser.parse_model_structure(raw_structure)
        if not parsed_model or "fields" not in parsed_model:
            logger.warning(f"Could not parse model structure: {model_name}")
            return None

        fields = [f for f in parsed_model["fields"] if f["name"] not in self.field_parser.metadata_fields]
        rows = []

        for field in fields:
            field_name = field["name"]
            type_display = self._format_type_display(field)
            description = field.get("description") or ""

            if field.get("related_model"):
                field_name = field_name[:-2] if field_name.endswith("Id") else field_name
                type_display += f" (FK → {field['related_model']})"
                description = f"Enter the primary identifier (e.g. name) of the {field['related_model']} record"

            rows.append(
                {
                    "field_name": field_name,
                    "type_display": type_display,
                    "is_required": field["is_required"],
                    "description": description,
                }
            )

        return rows

    def _get_enum_display_names(self, enums: Dict[str, List[str]], known_prefixes: Set[str]) -> Dict[str, str]:
        """Map each GraphQL enum name to a user-friendly display name.

        Groups enums with identical values — these represent the same concept under different
        Amplify-generated names (e.g. SpeciesRedListGlobal + SpeciesRedListMed when both use
        ConservationStatus). Picks the best name per group: prefers names not prefixed by a
        known model/type name (i.e. canonical standalone names), then shortest stripped name.
        """

        def strip_prefix(name: str) -> str:
            for prefix in sorted(known_prefixes, key=len, reverse=True):
                if name.startswith(prefix) and len(name) > len(prefix):
                    return name[len(prefix) :]
            return name

        stripped = {gql_name: strip_prefix(gql_name) for gql_name in enums}

        # Group by sorted values — enums with same values are the same concept
        by_values: Dict[tuple, List[str]] = {}
        for gql_name, values in enums.items():
            by_values.setdefault(tuple(sorted(values)), []).append(gql_name)

        display_names: Dict[str, str] = {}
        for value_key, gql_names in by_values.items():
            if len(gql_names) == 1:
                # Solo enum: stripping a prefix provides no disambiguation benefit and can lose
                # context (e.g. MediaQuality → Quality). Keep the original name.
                display_names[gql_names[0]] = gql_names[0]
            else:
                # Multiple enums share the same values — find the best canonical name.
                # Prefer a name unchanged by prefix-stripping (standalone canonical name).
                unchanged = [n for n in gql_names if stripped[n] == n]
                best = min(unchanged, key=len) if unchanged else min((stripped[n] for n in gql_names), key=len)
                for gql_name in gql_names:
                    display_names[gql_name] = best

        # Detect display-name conflicts: same display name, different value groups
        by_display: Dict[str, Set[tuple]] = {}
        for gql_name, display_name in display_names.items():
            value_key = tuple(sorted(enums[gql_name]))
            by_display.setdefault(display_name, set()).add(value_key)

        for display_name, value_groups in by_display.items():
            if len(value_groups) > 1:
                # Two different value-groups ended up with the same display name — revert both
                for gql_name in list(display_names.keys()):
                    if display_names[gql_name] == display_name:
                        display_names[gql_name] = gql_name

        return display_names

    @staticmethod
    def _build_display_enums(enums: Dict[str, List[str]], display_names: Dict[str, str]) -> Dict[str, List[str]]:
        """Build a deduplicated enum dict keyed by display name."""
        display_enums: Dict[str, List[str]] = {}
        for gql_name, values in enums.items():
            dn = display_names[gql_name]
            display_enums.setdefault(dn, values)
        return display_enums

    @staticmethod
    def _apply_display_names(text: str, display_names: Dict[str, str]) -> str:
        """Replace backtick-quoted GraphQL enum names with their display names."""
        for gql_name, display_name in sorted(display_names.items(), key=lambda x: len(x[0]), reverse=True):
            if gql_name != display_name:
                text = text.replace(f"`{gql_name}`", f"`{display_name}`")
        return text

    @staticmethod
    def _is_user_defined_type(type_name: str) -> bool:
        if type_name in _AMPLIFY_SYSTEM_TYPES:
            return False
        if any(type_name.startswith(p) for p in _AMPLIFY_GENERATED_PREFIXES):
            return False
        if any(type_name.endswith(s) for s in _AMPLIFY_GENERATED_SUFFIXES):
            return False
        return True

    @staticmethod
    def _truncate_sheet_name(name: str) -> str:
        return name[:31]

    @staticmethod
    def _format_type_display(field: Dict[str, Any]) -> str:
        base_type = field["type"]
        type_display = f"`{base_type}`"

        if field["is_list"]:
            type_display = f"`[{base_type}]`"

        if field["is_enum"]:
            type_display += " (Enum)"
        elif field["is_custom_type"]:
            type_display += " (Custom Type)"

        return type_display

    def _get_enum_values(self, enum_name: str) -> List[str]:
        enum_structure = self.client.get_model_structure(enum_name)
        if not enum_structure or "enumValues" not in enum_structure:
            return []
        return [ev["name"] for ev in enum_structure["enumValues"]]

    def _get_custom_type_fields(self, type_name: str) -> List[Dict[str, Any]]:
        type_structure = self.client.get_model_structure(type_name)
        if not type_structure:
            return []

        parsed = self.field_parser.parse_model_structure(type_structure)
        if not parsed or "fields" not in parsed:
            return []

        return [f for f in parsed["fields"] if f["name"] not in self.field_parser.metadata_fields]
